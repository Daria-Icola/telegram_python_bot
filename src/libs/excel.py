import numpy as np
import pandas as pd
import os
from openpyxl import load_workbook
import datetime
from src.bot.actions.get_reports import dict_for_detail_report, dict_for_final_report
from src import config

logger = config.logger

class Excel:
    def __init__(self):
        pass
        #self.workbook = workbook

    def find_by_number(self, number):
        wb = load_workbook(filename='./xls/cars.xlsx')
        sheet = wb['ПРОБЕГ']
        for row in sheet:
            for col in row:
                try:
                    if number in col.value:
                        return str(col.coordinate[0]) + str(self.find_by_rowname('Пробег'))
                except Exception as error:
                    logger.error(error)

    def find_by_rowname(self, name):
        wb = load_workbook(filename='xls/cars.xlsx')
        sheet = wb['ПРОБЕГ']
        for row in sheet:
            if name in row[0].value:
                try:
                    return str(row[0].coordinate)[1:]
                except Exception as error:
                    logger.error(error)

    def set_car_km(self):
        filename = os.getcwd() + '/xls/cars.xlsx'

        cars_info = dict_for_final_report()
        wb = load_workbook(filename)
        sheet = wb['ПРОБЕГ']
        new_path = os.getcwd() + '/reports'
        date = datetime.datetime.today().strftime('%d.%m.%Y')
        new_file_name = new_path + '/cars' + '.' + date + '.' + 'xls'
        for car in cars_info:
            try:
                car_number = self.find_by_number(car.get('number'))
                sheet[car_number].value = car.get('km_end')
                logger.info(f"Write in excel: {car.get('km_end')}")
            except Exception as error:
                logger.error("set_car_km ERROR: ", error)

        wb.save(new_file_name)
        return new_file_name

    def create_detail_report(self):
        filename = os.getcwd() + '/xls/reports.xlsx'
        rides_dict = dict_for_detail_report()
        date = datetime.datetime.today().strftime('%d.%m.%Y')
        name_excel_table = os.getcwd() + '/reports/' + 'сars.report.' + date + '.xlsx'
        logger.info('name_excel_table: ', name_excel_table)
        wb = load_workbook(filename)
        ws = wb.active

        for ride in rides_dict:
            try:
                row = [ride.get('ride_date'), ride.get('fullname'), ride.get('car_category'),
                       ride.get('car_brand'), ride.get('car_model'), ride.get('car_number'), ride.get('km_start'),
                       ride.get('km_end')]
                logger.info(row)
                ws.append(row)
            except Exception as error:
                logger.error(error)

        wb.save(name_excel_table)
        return name_excel_table
